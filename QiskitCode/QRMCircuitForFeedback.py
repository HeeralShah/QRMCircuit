from cmath import sqrt
from qiskit import *
from openpyxl import Workbook, load_workbook

#main imports for qiskit and general maths
import numpy as np
import math
from qiskit import QuantumCircuit, QuantumRegister, ClassicalRegister, transpile
from qiskit.providers.aer import QasmSimulator, AerSimulator
from qiskit.visualization import plot_histogram
from qiskit.extensions import UnitaryGate
from qiskit.circuit.library import CXGate, CZGate, XGate, ZGate, HGate, IGate, Measure
from qiskit.quantum_info.operators import Operator
from qiskit.circuit.exceptions import CircuitError


#noise packages
import qiskit.providers.aer.noise as noise

#time package to estimate computation sizes
from datetime import datetime


#setting self-controlled constant value
PARAM = 0.032
RUNS = 100000
PRE_CORRECT = True
POST_CORRECT = True
SYND_REP = False
ERR_CORR_NOISE = False

#for checking syndrome pair 2 only
chk_synd2 = False

syndrome_counter = 0
zero_counter = 0
one_counter = 0

#creating custom gates for adding depolarizing noise

#custom gates which will not have noise added

noiseless_meas = Measure()

x_gate = XGate()
z_gate = ZGate()


cx_gate = CXGate()
cz_gate = CZGate()

h_gate = HGate()

#custom gates which will have noise added

noisy_meas = Measure()

x_noisy = XGate(label="x_noisy")
z_noisy = ZGate(label="z_noisy")


cx_noisy = CXGate(label="cx_noise")
cz_noisy = CZGate(label="cz_noise")

h_noisy = HGate(label="h_noise")

x_unitary = UnitaryGate([[0,1],[1,0]])
z_unitary = UnitaryGate([[1,0],[0,-1]])

cx_unitary = UnitaryGate([[1,0,0,0]
                          [0,1,0,0],
                          [0,0,0,1]
                          [0,0,1,0]])

cz_unitary = UnitaryGate([[1,0,0,0]
                          [0,1,0,0],
                          [0,0,1,0]
                          [0,0,0,-1]])

h_unitary = UnitaryGate((1/np.sqrt(2)) * [[1,1],[1,-1]])

###########################INITIALIZE BIT AND QUBIT REGISTERS############################################


# Create a Quantum Circuit acting on the q register and adding classical bit for final measurement
circuit = QuantumCircuit()
main = QuantumRegister(15, 'q')
circuit.add_register(main)
circuit.qregs


final = ClassicalRegister(1, 'out')


#x and z stabilizer code generators
x_gens = [[0,2,4,6,8,10,12,14],
          [1,2,5,6,9,10,13,14],
          [3,4,5,6,11,12,13,14],
          [7,8,9,10,11,12,13,14]]

z_gens = [[0,2,4,6,8,10,12,14],
          [1,2,5,6,9,10,13,14],
          [3,4,5,6,11,12,13,14],
          [7,8,9,10,11,12,13,14]]


##initializing registers for syndrome measurement

qr = QuantumRegister(8, 'anc')
cr = ClassicalRegister(8, 'c')
check1 = QuantumRegister(1, "c1")
check2 = QuantumRegister(1, "c2")
check3 = QuantumRegister(1, "c3")
check4 = QuantumRegister(1, "c4")
meas1 = ClassicalRegister(1, "m1")
meas2 = ClassicalRegister(1, "m2")
meas3 = ClassicalRegister(1, "m3")
meas4 = ClassicalRegister(1, "m4")

circuit.add_register(qr)
circuit.add_register(cr)
circuit.add_register(check1)
circuit.add_register(check2)
circuit.add_register(check3)
circuit.add_register(check4)
circuit.add_register(meas1)
circuit.add_register(meas2)
circuit.add_register(meas3)
circuit.add_register(meas4)


circuit.qregs
circuit.cregs


########################################################################################

#functions for circuit construction

def qrm_encoding():
    #fixing initial state
    plus_states = [0,1,2,3,7]

    for i in plus_states:
        circuit.append(h_gate, [main[i]])
  
    
    #encoding state with QRM
    ctrl_qb = [[0,1,2,3,7], [0,1,2,3,7], [0,1,2,3,7], [0,1,7,2,3], [0,7,1,2,3], [0,7,2,3], [3,7,1]]
    targ_qb = [[11,13,14,12,10], [12,10,11,13,14], [10,11,9,14,13], [9,6,12,8,11], [6,11,8,5,4], [5,9,4,6], [5,8,4]]

    i=0
    while i < len(ctrl_qb):
        circuit.barrier()
        j=0
        while j < len(ctrl_qb[i]):
            circuit.append(cx_gate, [ctrl_qb[i][j], targ_qb[i][j]])
            #circuit.cx(ctrl_qb[i][j],targ_qb[i][j])
            j+=1
        i+=1


def measure_decode():
    #measurement circuit for final qubit
    circuit.add_register(final)
    circuit.cregs
    
    circuit.barrier()
    for i in range(0,15):
        circuit.s(i)
    circuit.barrier()
        
    
    #decoding circuit
    ctrl_qb = [[0,1,2,3,7], [0,1,2,3,7], [0,1,2,3,7], [0,1,7,2,3], [0,7,1,2,3], [0,7,2,3], [3,7,1]]
    targ_qb = [[11,13,14,12,10], [12,10,11,13,14], [10,11,9,14,13], [9,6,12,8,11], [6,11,8,5,4], [5,9,4,6], [5,8,4]]
    i=len(ctrl_qb)-1
    while i > -1:
        j=len(ctrl_qb[i])-1
        while j > -1:
            circuit.append(cx_gate, [ctrl_qb[i][j], targ_qb[i][j]])
            #circuit.cx(ctrl_qb[i][j],targ_qb[i][j])
            j-=1
        circuit.barrier()
        i-=1
        
        
    circuit.append(h_gate, [0])
    circuit.append(noiseless_meas, [main[0]], [final[0]])
    
    for i in range (0,8):
        circuit.reset(qr[i])
        circuit.append(noiseless_meas, [qr[i]], [cr[i]])
        

def logical_s():
    #applying logical S phase gate (which here is transversal S dagger)
    circuit.barrier()
    for i in range (0,15):
        circuit.sdg(i)
    circuit.barrier()

    
def add_err(x_err, z_err):  
    #add bit or phase flip
    if x_err > -1:
        circuit.x(main[x_err])
    if z_err > -1:
        circuit.z(main[z_err])


###syndrome circuit with stabilizer generators
def build_syndrome(generators, err_type):

    qc = QuantumCircuit()

    main = QuantumRegister(15)
    qc.add_register(main)
    qc.qregs

    qr = QuantumRegister(8)
    qc.add_register(qr)
    qc.qregs

    measurements = int(getattr(qr, "size") / 2)


    for i in range (0, measurements * 2):
        qc.append(h_noisy, [qr[i]])
    

    if err_type == "x_err":
        for i in range(0, measurements):
           for x in generators[i]:
               qc.append(cz_noisy, [qr[i], main[x]])
               qc.append(cz_noisy, [qr[i + measurements], main[x]])
    elif err_type == "z_err":
        for i in range(0, measurements):
            for x in generators[i]:
                qc.append(cx_noisy, [qr[i], main[x]])
                qc.append(cx_noisy, [qr[i + measurements], main[x]])
       
    for i in range (0, measurements * 2):
        qc.append(h_noisy, [qr[i]])
    
    gate = qc.to_instruction()
    return gate


### performing error correction

def apply_fix(generators, error_type, check, meas):

    register_size = int(getattr(main, "size"))
    measurements = 4 
    
    circuit.append(build_syndrome(generators, error_type), main[:] +  qr[:])
    for i in range(0, measurements * 2):
        circuit.append(noisy_meas, [qr[i]], [cr[i]])
    for i in range(0, measurements * 2):
        circuit.reset(qr[i])

    #circuit.append(x_gate, [check[0]]).c_if(cr, 0) 
    circuit.x(check[0]).c_if(cr,0) #for flipping meas val to one to NOT activate second syndrome
    for i in range(0, register_size):
        if error_type == "x_err":
            #circuit.append(x_noisy, [main[i]]).c_if(cr, (i+1) * 17)
            circuit.x(main[i]).c_if(cr, (i+1) * 17)
        elif error_type == "z_err":
            #circuit.append(z_noisy, [main[i]]).c_if(cr, (i+1) * 17)
            circuit.z(main[i]).c_if(cr, (i+1) * 17)
        #circuit.append(x_gate, [check[0]]).c_if(cr, (i+1) * 17) #for flipping meas val to one to NOT activate second syndrome
        circuit.x(check[0]).c_if(cr, (i+1) * 17)
    
    #####################################################################
    #for isolating tests of the second syndrome pair
    if chk_synd2 == True:
        circuit.reset(check[0]) 
        circuit.append(x_gate, [check[0]])

    #####################################################################
    
    circuit.append(noiseless_meas, [check[0]], [meas[0]])


    if SYND_REP == True:
        circuit.append(build_syndrome(generators, error_type), main[:] +  qr[:])
        for i in range(0, measurements * 2):
            circuit.append(noisy_meas, [qr[i]], [cr[i]])
        for i in range(0, measurements * 2):
            circuit.reset(qr[i])


#lines with c_if(meas, 0) should only execute if first pair of syndrome measurements are inconsistent
        circuit.x(check[0]).c_if(cr, 0).c_if(meas, 0)
        for i in range(0, register_size):
            if error_type == "x_err":
                #circuit.append(x_noisy, [main[i]]).c_if(cr, (i+1) * 17).c_if(meas, 0)
                circuit.x(main[i]).c_if(cr, (i+1) * 17).c_if(meas, 0) 
            elif error_type == "z_err":
                #circuit.append(z_noisy, [main[i]]).c_if(cr, (i+1) * 17).c_if(meas, 0)
                circuit.z(main[i]).c_if(cr, (i+1) * 17).c_if(meas, 0)
            #circuit.append(x_gate, [check[0]]).c_if(cr, (i+1) * 17).c_if(meas, 0)
            circuit.x(check[0]).c_if(meas, 0).c_if(cr, (i+1) * 17)
    
    
        circuit.append(noiseless_meas, [check[0]], [meas[0]])


###################################################################################################

#CREATNG DEPOLARIZING NOISE MODEL
noise_model = noise.NoiseModel()

depol_param = PARAM
error1 = noise.depolarizing_error(depol_param, 1)
error2 = noise.depolarizing_error(depol_param, 2)
phase_flip = noise.pauli_error([('Z', PARAM), ('I', 1 - PARAM)])

noise_model.add_all_qubit_quantum_error(error1, ['sdg'])

#noise_model.add_quantum_error(phase_flip, ["sdg"], [10])
#noise_model.add_quantum_error(phase_flip, ["sdg"], [9])

if ERR_CORR_NOISE == True:
    noise_model.add_all_qubit_quantum_error(error1, [x_noisy.label, z_noisy.label, h_noisy.label])
    noise_model.add_all_qubit_quantum_error(error2, [cx_noisy.label, cz_noisy.label])

noise_model.add_basis_gates(['sdg', "x", "z", "cx", "cz", "h", 'measure'])



######################BUILDING FINAL CIRCUIT####################

qrm_encoding()

if PRE_CORRECT == True:
    apply_fix(z_gens, "x_err", check1, meas1)
    apply_fix(x_gens, "z_err", check2, meas2)

logical_s()


if POST_CORRECT == True:
    apply_fix(z_gens, "x_err", check3, meas3)
    apply_fix(x_gens, "z_err", check4, meas4)

measure_decode()


#################################################################

def make_excel(filename):

    #time for spreadsheet
    START = datetime.now()

#################################################################

    result = execute(circuit, Aer.get_backend('qasm_simulator'),
                     noise_model=noise_model, shots=RUNS).result()
    hist = result.get_counts(0)
    print(hist)

    if PRE_CORRECT == True:
        SUCCESS = hist["0 1 1 1 1 00000000"]
        FAILURE = hist["1 1 1 1 1 00000000"]
    elif PRE_CORRECT == False and POST_CORRECT == True:
        SUCCESS = hist["0 1 1 0 0 00000000"]
        FAILURE = hist["1 1 1 0 0 00000000"]
    elif PRE_CORRECT == False and POST_CORRECT == False:
        SUCCESS = hist["0 0 0 0 0 00000000"]
        FAILURE = hist["1 0 0 0 0 00000000"]
#################################################################




    #times for spreadsheet
    FINISH = datetime.now()
    DIFF = FINISH - START


    #label for simulation type for spreadsheet

    SIM_TYPE = ""
    if PRE_CORRECT == True and POST_CORRECT == True:
        SIM_TYPE = "full circuit"
    elif PRE_CORRECT == False and POST_CORRECT == True:
        SIM_TYPE = "post-correct only"
    elif PRE_CORRECT == False and POST_CORRECT == False:
        SIM_TYPE = "logical S only"


        

    # Start by opening the spreadsheet and selecting the main sheet

    try:

        workbook = load_workbook("circuit_results.xlsx")
        worksheet = workbook["new_data"]

        new_data = [PARAM, RUNS, SUCCESS, FAILURE, START, FINISH, DIFF, ERR_CORR_NOISE, SYND_REP, SIM_TYPE]

        worksheet.append(new_data)

        workbook.save("circuit_results.xlsx")

    except PermissionError:

        open_workbook = Workbook()
        worksheet = open_workbook.active

        new_data = [PARAM, RUNS, FAILURE, START, FINISH, DIFF, SIM_TYPE]

        worksheet.append(new_data)

        open_workbook.save(filename + ".xlsx")




make_excel("new_circ")
